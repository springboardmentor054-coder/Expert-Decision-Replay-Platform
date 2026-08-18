from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from sqlalchemy import func, text

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.database.connection import SessionLocal

from app.models.decision import Decision
from app.models.approval import Approval
from app.models.user import User
from app.models.audit_log import AuditLog
from app.models.team import Team


router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


# ==========================================
# DATABASE DEPENDENCY
# ==========================================

def get_db():

    db = SessionLocal()

    try:
        yield db

    finally:
        db.close()


# ==========================================
# DECISION REPORT STATUS DEFINITIONS
# ==========================================

REPORT_STATUSES = [
    "Approved",
    "Rejected",
    "Pending",
    "Under Review",
    "In Review",
    "Review"
]


PENDING_STATUSES = [
    "Pending",
    "Under Review",
    "In Review",
    "Review"
]


# ==========================================
# SHARED REPORT DATA FUNCTIONS
# ==========================================

def build_decision_report(db: Session):

    # ------------------------------------------
    # LOAD CATEGORY NAMES FROM EXISTING TABLE
    # ------------------------------------------

    category_rows = db.execute(
        text(
            "SELECT id, name FROM categories"
        )
    ).fetchall()

    category_map = {
        row[0]: row[1]
        for row in category_rows
    }

    # ------------------------------------------
    # TOTAL DECISIONS
    #
    # Only Approved, Rejected and Pending-family
    # statuses are included.
    #
    # "done" is intentionally ignored.
    # ------------------------------------------

    total_decisions = db.query(
        func.count(Decision.id)
    ).filter(
        Decision.status.in_(REPORT_STATUSES)
    ).scalar() or 0

    # ------------------------------------------
    # APPROVED DECISIONS
    # ------------------------------------------

    approved_decisions = db.query(
        func.count(Decision.id)
    ).filter(
        Decision.status == "Approved"
    ).scalar() or 0

    # ------------------------------------------
    # REJECTED DECISIONS
    # ------------------------------------------

    rejected_decisions = db.query(
        func.count(Decision.id)
    ).filter(
        Decision.status == "Rejected"
    ).scalar() or 0

    # ------------------------------------------
    # PENDING DECISIONS
    #
    # Pending
    # Under Review
    # In Review
    # Review
    #
    # are all displayed as "Pending".
    # ------------------------------------------

    pending_decisions = db.query(
        func.count(Decision.id)
    ).filter(
        Decision.status.in_(PENDING_STATUSES)
    ).scalar() or 0

    # ------------------------------------------
    # FETCH REPORT DECISIONS
    # ------------------------------------------

    decisions = (
        db.query(
            Decision,
            User.name
        )
        .join(
            User,
            Decision.created_by == User.id
        )
        .filter(
            Decision.status.in_(REPORT_STATUSES)
        )
        .order_by(
            Decision.created_at.desc()
        )
        .all()
    )

    decision_details = []

    for decision, creator_name in decisions:

        # --------------------------------------
        # NORMALIZE STATUS FOR REPORT
        # --------------------------------------

        if decision.status in PENDING_STATUSES:

            report_status = "Pending"

        elif decision.status == "Approved":

            report_status = "Approved"

        elif decision.status == "Rejected":

            report_status = "Rejected"

        else:

            continue

        # --------------------------------------
        # GET CATEGORY NAME
        # --------------------------------------

        if decision.category_id is not None:

            category_name = category_map.get(
                decision.category_id,
                "Uncategorized"
            )

        else:

            category_name = "Uncategorized"

        # --------------------------------------
        # ADD DECISION TO REPORT
        # --------------------------------------

        decision_details.append({
            "title": decision.title,
            "category": category_name,
            "status": report_status,
            "created_by": creator_name,
            "created_date": decision.created_at
        })

    return {
        "report": "Decision Report",

        "summary": {
            "total_decisions": total_decisions,
            "approved": approved_decisions,
            "rejected": rejected_decisions,
            "pending": pending_decisions
        },

        "decisions": decision_details
    }


def build_approval_report(db: Session):

    reviewer_summary = (
        db.query(
            User.name.label("reviewer_name"),

            func.count(
                func.nullif(
                    Approval.status != "Approved",
                    True
                )
            ).label("decisions_approved"),

            func.count(
                func.nullif(
                    Approval.status != "Rejected",
                    True
                )
            ).label("decisions_rejected"),

            func.count(
                func.nullif(
                    Approval.status != "Pending",
                    True
                )
            ).label("pending_reviews")
        )
        .join(
            User,
            Approval.reviewer_id == User.id
        )
        .group_by(
            User.id,
            User.name
        )
        .order_by(
            User.name
        )
        .all()
    )

    return {
        "report": "Approval Report",

        "reviewers": [
            {
                "reviewer_name": reviewer_name,
                "decisions_approved": decisions_approved,
                "decisions_rejected": decisions_rejected,
                "pending_reviews": pending_reviews
            }
            for (
                reviewer_name,
                decisions_approved,
                decisions_rejected,
                pending_reviews
            ) in reviewer_summary
        ]
    }


def build_team_report(db: Session):

    teams = db.query(Team).order_by(
        Team.name
    ).all()

    team_reports = []

    for team in teams:

        total_users = db.query(
            func.count(User.id)
        ).filter(
            User.team_id == team.id
        ).scalar()

        total_decisions = (
            db.query(
                func.count(Decision.id)
            )
            .join(
                User,
                Decision.created_by == User.id
            )
            .filter(
                User.team_id == team.id
            )
            .scalar()
        )

        total_approvals = (
            db.query(
                func.count(Approval.id)
            )
            .join(
                User,
                Approval.reviewer_id == User.id
            )
            .filter(
                User.team_id == team.id
            )
            .scalar()
        )

        team_reports.append({
            "team_name": team.name,
            "total_decisions": total_decisions,
            "total_users": total_users,
            "total_approvals": total_approvals
        })

    return {
        "report": "Team Report",
        "teams": team_reports
    }


def build_audit_report(db: Session):

    total_logins = db.query(
        func.count(AuditLog.id)
    ).filter(
        AuditLog.action_type == "LOGIN"
    ).scalar()

    decisions_created = db.query(
        func.count(AuditLog.id)
    ).filter(
        AuditLog.action_type == "DECISION_CREATED"
    ).scalar()

    documents_uploaded = db.query(
        func.count(AuditLog.id)
    ).filter(
        AuditLog.action_type == "DOCUMENT_UPLOADED"
    ).scalar()

    comments_added = db.query(
        func.count(AuditLog.id)
    ).filter(
        AuditLog.action_type == "COMMENT_ADDED"
    ).scalar()

    approval_actions = db.query(
        func.count(AuditLog.id)
    ).filter(
        AuditLog.action_type.in_([
            "DECISION_APPROVED",
            "DECISION_REJECTED"
        ])
    ).scalar()

    return {
        "report": "Audit Report",
        "total_logins": total_logins,
        "decisions_created": decisions_created,
        "documents_uploaded": documents_uploaded,
        "comments_added": comments_added,
        "approval_actions": approval_actions
    }


# ==========================================
# TASK 2 - DECISION REPORT
# ==========================================

@router.get("/decisions")
def get_decision_report(
    db: Session = Depends(get_db)
):

    return build_decision_report(db)


# ==========================================
# TASK 3 - APPROVAL REPORT
# ==========================================

@router.get("/approvals")
def get_approval_report(
    db: Session = Depends(get_db)
):

    return build_approval_report(db)


# ==========================================
# TASK 4 - TEAM REPORT
# ==========================================

@router.get("/teams")
def get_team_report(
    db: Session = Depends(get_db)
):

    return build_team_report(db)


# ==========================================
# TASK 5 - AUDIT REPORT
# ==========================================

@router.get("/audit")
def get_audit_report(
    db: Session = Depends(get_db)
):

    return build_audit_report(db)


# ==========================================
# TASK 6 - EXPORT PDF
# ==========================================

@router.get("/export/pdf")
def export_reports_pdf(
    db: Session = Depends(get_db)
):

    # ------------------------------------------
    # GET CURRENT LIVE REPORT DATA
    # ------------------------------------------

    decision_report = build_decision_report(db)
    approval_report = build_approval_report(db)
    team_report = build_team_report(db)
    audit_report = build_audit_report(db)

    # ------------------------------------------
    # CREATE PDF IN MEMORY
    # ------------------------------------------

    pdf_buffer = BytesIO()

    document = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    title_style.alignment = TA_CENTER

    heading_style = styles["Heading2"]

    normal_style = styles["BodyText"]

    elements = []

    # ------------------------------------------
    # TITLE
    # ------------------------------------------

    elements.append(
        Paragraph(
            "Expert Decision Replay Platform",
            title_style
        )
    )

    elements.append(
        Paragraph(
            "Reports Export",
            heading_style
        )
    )

    elements.append(Spacer(1, 20))

    # ==========================================
    # DECISION REPORT
    # ==========================================

    elements.append(
        Paragraph(
            "Decision Report",
            heading_style
        )
    )

    decision_summary = decision_report["summary"]

    decision_summary_table = Table([
        [
            "Total Decisions",
            "Approved",
            "Rejected",
            "Pending"
        ],
        [
            decision_summary["total_decisions"],
            decision_summary["approved"],
            decision_summary["rejected"],
            decision_summary["pending"]
        ]
    ])

    decision_summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
        ])
    )

    elements.append(decision_summary_table)
    elements.append(Spacer(1, 15))

    decision_rows = [
        [
            "Title",
            "Category",
            "Status",
            "Created By",
            "Created Date"
        ]
    ]

    for decision in decision_report["decisions"]:

        created_date = decision["created_date"]

        if created_date:
            created_date = created_date.strftime(
                "%d/%m/%Y %H:%M"
            )

        decision_rows.append([
            str(decision["title"]),
            str(decision["category"]),
            str(decision["status"]),
            str(decision["created_by"]),
            str(created_date)
        ])

    if len(decision_rows) > 1:

        decision_table = Table(
            decision_rows,
            repeatRows=1
        )

        decision_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
            ])
        )

        elements.append(decision_table)

    elements.append(Spacer(1, 25))

    # ==========================================
    # APPROVAL REPORT
    # ==========================================

    elements.append(
        Paragraph(
            "Approval Report",
            heading_style
        )
    )

    approval_rows = [
        [
            "Reviewer",
            "Approved",
            "Rejected",
            "Pending"
        ]
    ]

    for reviewer in approval_report["reviewers"]:

        approval_rows.append([
            str(reviewer["reviewer_name"]),
            reviewer["decisions_approved"],
            reviewer["decisions_rejected"],
            reviewer["pending_reviews"]
        ])

    if len(approval_rows) > 1:

        approval_table = Table(
            approval_rows,
            repeatRows=1
        )

        approval_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
            ])
        )

        elements.append(approval_table)

    elements.append(Spacer(1, 25))

    # ==========================================
    # TEAM REPORT
    # ==========================================

    elements.append(
        Paragraph(
            "Team Report",
            heading_style
        )
    )

    team_rows = [
        [
            "Team",
            "Total Users",
            "Total Decisions",
            "Total Approvals"
        ]
    ]

    for team in team_report["teams"]:

        team_rows.append([
            str(team["team_name"]),
            team["total_users"],
            team["total_decisions"],
            team["total_approvals"]
        ])

    if len(team_rows) > 1:

        team_table = Table(
            team_rows,
            repeatRows=1
        )

        team_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
            ])
        )

        elements.append(team_table)

    elements.append(Spacer(1, 25))

    # ==========================================
    # AUDIT REPORT
    # ==========================================

    elements.append(
        Paragraph(
            "Audit Report",
            heading_style
        )
    )

    audit_rows = [
        ["Audit Action", "Count"],
        ["Total Logins", audit_report["total_logins"]],
        ["Decisions Created", audit_report["decisions_created"]],
        ["Documents Uploaded", audit_report["documents_uploaded"]],
        ["Comments Added", audit_report["comments_added"]],
        ["Approval Actions", audit_report["approval_actions"]]
    ]

    audit_table = Table(
        audit_rows,
        repeatRows=1
    )

    audit_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
        ])
    )

    elements.append(audit_table)

    # ------------------------------------------
    # BUILD PDF
    # ------------------------------------------

    document.build(elements)

    pdf_buffer.seek(0)

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                "attachment; filename=expert_decision_reports.pdf"
        }
    )


# ==========================================
# TASK 6 - EXPORT EXCEL
# ==========================================

@router.get("/export/excel")
def export_reports_excel(
    db: Session = Depends(get_db)
):

    # ------------------------------------------
    # GET CURRENT LIVE REPORT DATA
    # ------------------------------------------

    decision_report = build_decision_report(db)
    approval_report = build_approval_report(db)
    team_report = build_team_report(db)
    audit_report = build_audit_report(db)

    # ------------------------------------------
    # CREATE WORKBOOK
    # ------------------------------------------

    workbook = Workbook()

    # ==========================================
    # DECISION REPORT SHEET
    # ==========================================

    decision_sheet = workbook.active
    decision_sheet.title = "Decision Report"

    decision_sheet.append([
        "Decision Report"
    ])

    decision_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    decision_sheet.append([])

    decision_summary = decision_report["summary"]

    decision_sheet.append([
        "Total Decisions",
        "Approved",
        "Rejected",
        "Pending"
    ])

    decision_sheet.append([
        decision_summary["total_decisions"],
        decision_summary["approved"],
        decision_summary["rejected"],
        decision_summary["pending"]
    ])

    decision_sheet.append([])

    decision_sheet.append([
        "Title",
        "Category",
        "Status",
        "Created By",
        "Created Date"
    ])

    for decision in decision_report["decisions"]:

        created_date = decision["created_date"]

        if created_date:
            created_date = created_date.strftime(
                "%d/%m/%Y %H:%M"
            )

        decision_sheet.append([
            decision["title"],
            decision["category"],
            decision["status"],
            decision["created_by"],
            created_date
        ])

    # ==========================================
    # APPROVAL REPORT SHEET
    # ==========================================

    approval_sheet = workbook.create_sheet(
        "Approval Report"
    )

    approval_sheet.append([
        "Approval Report"
    ])

    approval_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    approval_sheet.append([])

    approval_sheet.append([
        "Reviewer",
        "Approved",
        "Rejected",
        "Pending"
    ])

    for reviewer in approval_report["reviewers"]:

        approval_sheet.append([
            reviewer["reviewer_name"],
            reviewer["decisions_approved"],
            reviewer["decisions_rejected"],
            reviewer["pending_reviews"]
        ])

    # ==========================================
    # TEAM REPORT SHEET
    # ==========================================

    team_sheet = workbook.create_sheet(
        "Team Report"
    )

    team_sheet.append([
        "Team Report"
    ])

    team_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    team_sheet.append([])

    team_sheet.append([
        "Team",
        "Total Users",
        "Total Decisions",
        "Total Approvals"
    ])

    for team in team_report["teams"]:

        team_sheet.append([
            team["team_name"],
            team["total_users"],
            team["total_decisions"],
            team["total_approvals"]
        ])

    # ==========================================
    # AUDIT REPORT SHEET
    # ==========================================

    audit_sheet = workbook.create_sheet(
        "Audit Report"
    )

    audit_sheet.append([
        "Audit Report"
    ])

    audit_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    audit_sheet.append([])

    audit_sheet.append([
        "Audit Action",
        "Count"
    ])

    audit_sheet.append([
        "Total Logins",
        audit_report["total_logins"]
    ])

    audit_sheet.append([
        "Decisions Created",
        audit_report["decisions_created"]
    ])

    audit_sheet.append([
        "Documents Uploaded",
        audit_report["documents_uploaded"]
    ])

    audit_sheet.append([
        "Comments Added",
        audit_report["comments_added"]
    ])

    audit_sheet.append([
        "Approval Actions",
        audit_report["approval_actions"]
    ])

    # ==========================================
    # FORMAT ALL SHEETS
    # ==========================================

    for sheet in workbook.worksheets:

        for row in sheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center"
                )

        # Bold header rows
        for cell in sheet[3]:

            cell.font = Font(
                bold=True
            )

        # Auto-size columns
        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                50
            )

    # ------------------------------------------
    # SAVE EXCEL IN MEMORY
    # ------------------------------------------

    excel_buffer = BytesIO()

    workbook.save(excel_buffer)

    excel_buffer.seek(0)

    return StreamingResponse(
        excel_buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                "attachment; filename=expert_decision_reports.xlsx"
        }
    )